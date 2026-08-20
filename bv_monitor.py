"""
Bona Vacantia List Monitor — SMTP2GO + Styled Excel Report version
=====================================================================
Runs on a schedule (e.g. GitHub Actions). Downloads the latest BV
Unclaimed Estates list from gov.uk, compares it to the previously
saved version, builds a styled ON/OFF Excel report, and emails it
via SMTP2GO — but ONLY if something changed.

Test mode:
    python bv_monitor.py --test-email
Sends a simple test email using the same SMTP2GO config, skipping
the gov.uk download/compare/report steps entirely. Useful for
confirming SMTP auth and delivery work without waiting for a real
list change.
"""

import os
import re
import ssl
import sys
import smtplib
import datetime
import requests
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ─────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────

SMTP_HOST     = "mail.smtp2go.com"
SMTP_PORT_SSL = 465    # SSL
SMTP_PORT_TLS = 2525   # STARTTLS

# Secrets — pulled from environment variables (set as GitHub Actions
# repo/org secrets). Never hardcode these in the script.
SMTP_USERNAME = os.environ["SMTP_USERNAME"]
SMTP_PASSWORD = os.environ["SMTP_PASSWORD"]

# Email addresses — also environment-driven so they aren't tied to the
# code and can be changed without a commit. EMAIL_FROM falls back to a
# sensible default if not set; EMAIL_TO is required.
EMAIL_FROM = os.environ.get("EMAIL_FROM", "bv@family-wise.co.uk")
EMAIL_TO   = os.environ["EMAIL_TO"]

SCRIPT_DIR      = Path(__file__).parent
SAVED_LIST_PATH = SCRIPT_DIR / "bv_saved_list.csv"
OUTPUT_XLSX     = SCRIPT_DIR / "BV_Changes.xlsx"
LOG_FILE        = SCRIPT_DIR / "bv_monitor.log"

BV_PAGE_URL = "https://www.gov.uk/government/statistical-data-sets/unclaimed-estates-list"

# ─────────────────────────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────────────────────────

def log(msg):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")

# ─────────────────────────────────────────────────────────────
#  STEP 1 — Find and download the latest BV spreadsheet
# ─────────────────────────────────────────────────────────────

def find_download_url() -> str:
    headers = {"User-Agent": "Mozilla/5.0 (BV-Monitor/1.0)"}
    resp = requests.get(BV_PAGE_URL, headers=headers, timeout=30)
    resp.raise_for_status()

    urls = re.findall(
        r'https://assets\.publishing\.service\.gov\.uk[^"\']+\.(?:xlsx|csv)',
        resp.text
    )

    if not urls:
        raise RuntimeError(
            "Could not find a download link on the BV page. "
            "The page layout may have changed."
        )

    return urls[0]


def download_bv_list(url: str) -> pd.DataFrame:
    headers = {"User-Agent": "Mozilla/5.0 (BV-Monitor/1.0)"}
    resp = requests.get(url, headers=headers, timeout=60)
    resp.raise_for_status()

    ext = ".xlsx" if url.endswith(".xlsx") else ".csv"
    tmp_path = SCRIPT_DIR / f"_tmp_bv{ext}"

    with open(tmp_path, "wb") as f:
        f.write(resp.content)

    df = pd.read_excel(tmp_path) if ext == ".xlsx" else pd.read_csv(tmp_path)
    tmp_path.unlink()

    df['BV Reference'] = df['BV Reference'].astype(str).str.strip()
    return df

# ─────────────────────────────────────────────────────────────
#  STEP 2 — Compare with saved list
# ─────────────────────────────────────────────────────────────

def load_saved_list():
    if SAVED_LIST_PATH.exists():
        df = pd.read_csv(SAVED_LIST_PATH, dtype=str)
        df['BV Reference'] = df['BV Reference'].astype(str).str.strip()
        return df
    return None


def save_list(df: pd.DataFrame):
    df.to_csv(SAVED_LIST_PATH, index=False)


def compare(old_df: pd.DataFrame, new_df: pd.DataFrame):
    old_refs = set(old_df['BV Reference'])
    new_refs = set(new_df['BV Reference'])

    sort_col = 'Surname' if 'Surname' in new_df.columns else new_df.columns[0]

    on_df  = new_df[new_df['BV Reference'].isin(new_refs - old_refs)].sort_values(sort_col)
    off_df = old_df[old_df['BV Reference'].isin(old_refs - new_refs)].sort_values(
        sort_col if sort_col in old_df.columns else old_df.columns[0]
    )

    return on_df, off_df

# ─────────────────────────────────────────────────────────────
#  STEP 3 — Build styled Excel report
# ─────────────────────────────────────────────────────────────

def style_sheet(ws, df, title, header_color, tab_color):
    ws.title = title
    ws.sheet_properties.tabColor = tab_color

    preferred = ['BV Reference', 'Forename', 'Surname', 'Date of Death', 'Place of Death', 'Notes', 'Other']
    for extra_col in ('Notes', 'Other'):
        if extra_col not in df.columns:
            df[extra_col] = ''
    cols = [c for c in preferred if c in df.columns] or list(df.columns)

    col_widths = {
        'BV Reference': 18, 'Forename': 15, 'Surname': 20,
        'Date of Death': 16, 'Place of Death': 35,
        'Notes': 30, 'Other': 20
    }

    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', start_color=header_color)

    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    ws['A1'] = title
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f'A2:{get_column_letter(len(cols))}2')
    ws['A2'] = f'Total entries: {len(df)}'
    ws['A2'].font      = Font(name='Arial', italic=True, size=10, color='555555')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=3, column=ci, value=col)
        cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 20)
    ws.row_dimensions[3].height = 18

    for ri, (_, row) in enumerate(df[cols].iterrows(), 4):
        fill_color = 'F9F9F9' if ri % 2 == 0 else 'FFFFFF'
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if hasattr(val, 'strftime'):
                val = val.strftime('%d/%m/%Y')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name='Arial', size=10)
            cell.fill      = PatternFill('solid', start_color=fill_color)
            cell.border    = border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A4'


def build_excel(on_df: pd.DataFrame, off_df: pd.DataFrame) -> Path:
    wb = Workbook()
    ws1 = wb.active
    style_sheet(ws1, on_df,  'ON List (New Entries)',      '2E7D32', '00AA00')
    ws2 = wb.create_sheet()
    style_sheet(ws2, off_df, 'OFF List (Removed Entries)', 'B71C1C', 'CC0000')
    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX

# ─────────────────────────────────────────────────────────────
#  STEP 4 — Build a nice HTML email body
# ─────────────────────────────────────────────────────────────

def build_html_body(on_df: pd.DataFrame, off_df: pd.DataFrame) -> str:
    today = datetime.date.today().strftime("%d %B %Y")

    def row_html(r, sign, color):
        return (
            f'<tr>'
            f'<td style="padding:6px 10px;color:{color};font-weight:bold;">{sign}</td>'
            f'<td style="padding:6px 10px;">{r.get("Forename","")} {r.get("Surname","")}</td>'
            f'<td style="padding:6px 10px;color:#666;">{r.get("BV Reference","")}</td>'
            f'</tr>'
        )

    on_rows = "".join(row_html(r, "+", "#2E7D32") for _, r in on_df.iterrows()) if not on_df.empty else \
        '<tr><td colspan="3" style="padding:8px 10px;color:#999;">No new entries.</td></tr>'

    off_rows = "".join(row_html(r, "–", "#B71C1C") for _, r in off_df.iterrows()) if not off_df.empty else \
        '<tr><td colspan="3" style="padding:8px 10px;color:#999;">No removed entries.</td></tr>'

    html = f"""\
<html>
  <body style="font-family:Arial, sans-serif; background-color:#f4f4f4; padding:20px; margin:0;">
    <div style="max-width:640px;margin:0 auto;background:#ffffff;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">

      <div style="background-color:#1a1a2e;padding:20px 24px;">
        <h1 style="color:#ffffff;margin:0;font-size:18px;">Bona Vacantia List Update</h1>
        <p style="color:#cfcfcf;margin:4px 0 0;font-size:13px;">{today}</p>
      </div>

      <div style="display:flex;padding:16px 24px 0 24px;">
        <div style="flex:1;background-color:#E8F5E9;border-radius:6px;padding:12px 16px;margin-right:8px;">
          <p style="margin:0;font-size:12px;color:#2E7D32;font-weight:bold;">NEW (ON)</p>
          <p style="margin:2px 0 0;font-size:22px;color:#2E7D32;font-weight:bold;">{len(on_df)}</p>
        </div>
        <div style="flex:1;background-color:#FFEBEE;border-radius:6px;padding:12px 16px;margin-left:8px;">
          <p style="margin:0;font-size:12px;color:#B71C1C;font-weight:bold;">REMOVED (OFF)</p>
          <p style="margin:2px 0 0;font-size:22px;color:#B71C1C;font-weight:bold;">{len(off_df)}</p>
        </div>
      </div>

      <div style="padding:20px 24px;">
        <h2 style="font-size:14px;color:#2E7D32;border-bottom:2px solid #2E7D32;padding-bottom:6px;">New Entries (ON)</h2>
        <table style="width:100%;border-collapse:collapse;font-size:13px;">
          {on_rows}
        </table>

        <h2 style="font-size:14px;color:#B71C1C;border-bottom:2px solid #B71C1C;padding-bottom:6px;margin-top:24px;">Removed Entries (OFF)</h2>
        <table style="width:100%;border-collapse:collapse;font-size:13px;">
          {off_rows}
        </table>
      </div>

      <div style="background-color:#f4f4f4;padding:14px 24px;border-top:1px solid #e0e0e0;">
        <p style="margin:0;font-size:12px;color:#777;">
          Full details are in the attached Excel report (<b>BV_Changes.xlsx</b>), with separate ON and OFF sheets.
        </p>
      </div>

    </div>
  </body>
</html>
"""
    return html


def build_plaintext_body(on_df: pd.DataFrame, off_df: pd.DataFrame) -> str:
    today = datetime.date.today().strftime("%d %B %Y")
    lines = [
        f"Bona Vacantia list changes detected — {today}.",
        "",
        f"NEW entries (ON list):      {len(on_df)}",
        f"REMOVED entries (OFF list): {len(off_df)}",
        "",
    ]
    if not on_df.empty:
        lines.append("New entries (ON):")
        for _, r in on_df.iterrows():
            lines.append(f"  + {r.get('Forename','')} {r.get('Surname','')}  [{r.get('BV Reference','')}]")
        lines.append("")
    if not off_df.empty:
        lines.append("Removed entries (OFF):")
        for _, r in off_df.iterrows():
            lines.append(f"  - {r.get('Forename','')} {r.get('Surname','')}  [{r.get('BV Reference','')}]")
        lines.append("")
    lines.append("Full details are in the attached Excel file.")
    return "\n".join(lines)

# ─────────────────────────────────────────────────────────────
#  STEP 5 — Send email via SMTP2GO (SSL 465, fallback TLS 2525)
# ─────────────────────────────────────────────────────────────

def send_email(on_df: pd.DataFrame, off_df: pd.DataFrame, xlsx_path: Path):
    today   = datetime.date.today().strftime("%d %B %Y")
    subject = f"Bona Vacantia List Update — {today} ({len(on_df)} new, {len(off_df)} removed)"

    msg            = MIMEMultipart("mixed")
    msg["From"]    = EMAIL_FROM
    msg["To"]      = EMAIL_TO
    msg["Subject"] = subject

    body_alt = MIMEMultipart("alternative")
    body_alt.attach(MIMEText(build_plaintext_body(on_df, off_df), "plain"))
    body_alt.attach(MIMEText(build_html_body(on_df, off_df), "html"))
    msg.attach(body_alt)

    with open(xlsx_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={xlsx_path.name}")
    msg.attach(part)

    recipients = [addr.strip() for addr in EMAIL_TO.split(",") if addr.strip()]

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT_SSL, context=context, timeout=30) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        log(f"Email sent to {EMAIL_TO} via SMTP2GO (SSL 465)")
    except Exception as ssl_err:
        log(f"SSL (465) send failed: {ssl_err}. Retrying with TLS (2525)...")
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT_TLS, timeout=30) as server:
            server.starttls(context=ssl.create_default_context())
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        log(f"Email sent to {EMAIL_TO} via SMTP2GO (TLS 2525)")


def send_test_email():
    """Sends a simple test email using the same SMTP2GO config and
    fallback logic as send_email(), without touching the gov.uk
    download or comparison logic. Run with: python bv_monitor.py --test-email
    """
    today   = datetime.date.today().strftime("%d %B %Y")
    subject = f"BV Monitor — Test Email ({today})"

    msg            = MIMEMultipart("alternative")
    msg["From"]    = EMAIL_FROM
    msg["To"]      = EMAIL_TO
    msg["Subject"] = subject

    plain = (
        "This is a test email from bv_monitor.py.\n\n"
        f"Sent: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"From: {EMAIL_FROM}\n"
        f"To:   {EMAIL_TO}\n\n"
        "If you're reading this, SMTP2GO authentication and delivery "
        "are both working correctly."
    )
    html = f"""\
<html><body style="font-family:Arial, sans-serif;">
  <h2 style="color:#2E7D32;">BV Monitor — Test Email ✅</h2>
  <p>This is a test email from <b>bv_monitor.py</b>.</p>
  <p style="color:#666;font-size:13px;">
    Sent: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
    From: {EMAIL_FROM}<br>
    To: {EMAIL_TO}
  </p>
  <p>If you're reading this, SMTP2GO authentication and delivery are both working correctly.</p>
</body></html>
"""
    msg.attach(MIMEText(plain, "plain"))
    msg.attach(MIMEText(html, "html"))

    recipients = [addr.strip() for addr in EMAIL_TO.split(",") if addr.strip()]

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT_SSL, context=context, timeout=30) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        log(f"✅ Test email sent to {EMAIL_TO} via SMTP2GO (SSL 465)")
    except Exception as ssl_err:
        log(f"SSL (465) test send failed: {ssl_err}. Retrying with TLS (2525)...")
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT_TLS, timeout=30) as server:
            server.starttls(context=ssl.create_default_context())
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        log(f"✅ Test email sent to {EMAIL_TO} via SMTP2GO (TLS 2525)")

# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────

def main():
    log("── BV Monitor starting ──")

    try:
        log("Finding download URL on gov.uk...")
        url = find_download_url()
        log(f"Downloading: {url}")
        new_df = download_bv_list(url)
        log(f"Downloaded {len(new_df)} entries.")

        old_df = load_saved_list()

        if old_df is None:
            log("No saved list found — saving current list as baseline. No email sent.")
            save_list(new_df)
            return

        on_df, off_df = compare(old_df, new_df)
        log(f"Comparison done — ON: {len(on_df)}, OFF: {len(off_df)}")

        if on_df.empty and off_df.empty:
            log("No changes detected. No email sent.")
            save_list(new_df)
            return

        log("Changes found — building Excel report...")
        xlsx = build_excel(on_df, off_df)

        log("Sending email...")
        send_email(on_df, off_df, xlsx)

        save_list(new_df)
        log("All done ✅")

    except Exception as e:
        log(f"ERROR: {e}")
        raise


if __name__ == "__main__":
    if "--test-email" in sys.argv:
        log("── Running in TEST EMAIL mode ──")
        send_test_email()
    else:
        main()

